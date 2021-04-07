'''
Created on 07-Apr-2021

@author: Ganga
'''
import pytest
import openpyxl
import time
import Utils.XLUtils
from selenium import webdriver
from Pages.test_LoginPage import LoginPage
from Pages.ChatPage import ChatPage
import os



browser = webdriver.Chrome(executable_path='C:/Users/g.siddalingappa/git/ctAssets/ct_assets_seleniumpythontestng/Drivers/chromedriver.exe')
browser.maximize_window()

login = LoginPage(driver=browser)     
chat = ChatPage(driver=browser)

@pytest.fixture
def login_chatbot(url, username, password):
    print ("Displaying url: %s" % url)
    print("Displaying username: %s" % username)    
    print("Displaying password: %s" % password)
    browser.get(url)
    login.enter_username(username)
    login.enter_password(password)
    Login_Text = login.get_login_btn_text()
    assert Login_Text == 'Login'
    login.click_login_button()
    
    yield 
    chat.click_accountmenu_sidebar()
    chat.click_logout()
    


def test_greetIntents(login_chatbot):
    wk = openpyxl.load_workbook("C:/Users/g.siddalingappa/git/excelData/AssetsTestDataNew.xlsx")
    path = "C:/Users/g.siddalingappa/git/excelData/AssetsTestDataNew.xlsx"
    rows = Utils.XLUtils.getRowCount(path, 'GreetIntents')   

    for i in range(2,rows+1): 
        test_input = Utils.XLUtils.readData(path, 'GreetIntents', i, 1)
        chat.enter_userinput(test_input)
        time.sleep(4)
        chat.get_userinput()
        time.sleep(9)
        chat.get_botoutput()
        bot_output = chat.get_botoutput()
        sh=wk["GreetIntents"]
        sh.cell(row=i, column=3).value = bot_output
        Utils.XLUtils.writeData(path, 'GreetIntents', i, 3, bot_output)
        expected_intent = Utils.XLUtils.readData(path, 'GreetIntents', i, 2)
        
        if expected_intent == bot_output:
            sh.cell(row=i, column=4).value = "Pass"
        else:
            sh.cell(row=i, column=4).value = "Fail"
                
    wk.save(path)


def test_clusterInfo(login_chatbot):
    wk = openpyxl.load_workbook("C:/Users/g.siddalingappa/git/excelData/AssetsTestDataNew.xlsx")
    path = "C:/Users/g.siddalingappa/git/excelData/AssetsTestDataNew.xlsx"
    rows = Utils.XLUtils.getRowCount(path, 'ClusterInfo') 
    
    for i in range(2,rows+1): 
        Test_input = Utils.XLUtils.readData(path, 'ClusterInfo', i, 1)
        chat.enter_userinput(Test_input)
        time.sleep(4)
        chat.get_userinput()
        time.sleep(9)
        chat.get_botoutput()
        bot_output = chat.get_botoutput()
        sh=wk["ClusterInfo"]
        sh.cell(row=i, column=3).value = bot_output
        Utils.XLUtils.writeData(path, 'ClusterInfo', i, 3, bot_output)    
    
    wk.save(path)
        
        
        


    
   
     
   


